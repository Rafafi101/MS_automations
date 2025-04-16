import os
import re
import sys
import logging
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment
from datetime import datetime, timedelta
import pytz
import xlrd
import pandas as pd


def configure_logging():
    logging.basicConfig(level=logging.DEBUG,
                        format='%(asctime)s %(levelname)s - %(message)s',
                        handlers=[
                            logging.FileHandler('compare_log.log'),
                            logging.StreamHandler(sys.stdout)
                        ])
    logging.info("Logging configured")


def get_file_location():
    if getattr(sys, 'frozen', False):
        file_location = os.path.dirname(sys.executable)
        logging.debug(
            f"Running in a bundled executable. File location: {file_location}")
    else:
        file_location = os.path.dirname(os.path.abspath(__file__))
        logging.debug(f"Running in a script. File location: {file_location}")
    return file_location


def convert_xls_to_xlsx(xls_path):
    logging.debug(f"Converting {xls_path} to .xlsx")
    with open(xls_path, 'r', encoding='utf-8') as file:
        content = file.read(1024)
        if content.strip().startswith('<html'):
            logging.debug(
                f"{xls_path} is an HTML file. Converting HTML to .xlsx")
            try:
                df = pd.read_html(xls_path)[0]
                xlsx_path = xls_path.replace('.xls', '.xlsx')
                with pd.ExcelWriter(xlsx_path, engine='openpyxl') as writer:
                    df.to_excel(writer, index=False, sheet_name='Exported')
                logging.info(f"Converted {xls_path} to {xlsx_path}")
                file.close()
                if 'Exported' not in load_workbook(xlsx_path).sheetnames:
                    logging.error(
                        f"Converted HTML file {xlsx_path} missing 'Exported' sheet.")
                    os.remove(xlsx_path)
                    return None
                os.remove(xls_path)
                return xlsx_path
            except Exception as e:
                logging.error(f"Failed to convert HTML file {xls_path}: {e}")
                return None
    try:
        xls_workbook = xlrd.open_workbook(xls_path)
    except xlrd.biffh.XLRDError as e:
        logging.error(f"Failed to open {xls_path}: {e}")
        return None
    xlsx_path = xls_path.replace('.xls', '.xlsx')
    xlsx_workbook = Workbook()
    xlsx_sheet = xlsx_workbook.create_sheet(title='Exported')
    for sheet_name in xls_workbook.sheet_names():
        xls_sheet = xls_workbook.sheet_by_name(sheet_name)
        for row_idx in range(xls_sheet.nrows):
            row = xls_sheet.row_values(row_idx)
            xlsx_sheet.append(row)
    if 'Sheet' in xlsx_workbook.sheetnames:
        del xlsx_workbook['Sheet']
    xlsx_workbook.save(xlsx_path)
    xls_workbook.release_resources()
    os.remove(xls_path)
    logging.info(f"Converted {xls_path} to {xlsx_path}")
    try:
        wb = load_workbook(xlsx_path)
        if 'Exported' not in wb.sheetnames:
            logging.error(
                f"Converted file {xlsx_path} missing 'Exported' sheet.")
            os.remove(xlsx_path)
            return None
        wb.close()
    except Exception as e:
        logging.error(f"Failed to verify 'Exported' sheet: {e}")
        return None
    return xlsx_path


def check_and_update_header(sheet):
    logging.debug(f"Checking and updating header for sheet: {sheet.title}")
    if sheet.cell(row=1, column=1).value != "Person Number":
        sheet.cell(row=1, column=1).value = "Person Number"
    for col in range(1, 10):
        column_letter = chr(64 + col)
        sheet.column_dimensions[column_letter].width = 20
        cell = sheet.cell(row=1, column=col)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
    sheet.auto_filter.ref = sheet.dimensions


def sort_by_column(sheet, column_name):
    column_index = None
    for col in range(1, sheet.max_column + 1):
        if sheet.cell(row=1, column=col).value == column_name:
            column_index = col
            break
    if column_index is None:
        raise Exception(
            f"Column '{column_name}' not found in sheet '{sheet.title}'")
    data = list(sheet.iter_rows(values_only=True))
    header = data[0]
    rows = data[1:]
    rows.sort(key=lambda x: (x[column_index - 1]
                             is not None, x[column_index - 1]))
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, max_col=sheet.max_column):
        for cell in row:
            cell.value = None
    for row_index, row_data in enumerate([header] + rows, start=1):
        for col_index, cell_value in enumerate(row_data, start=1):
            sheet.cell(row=row_index, column=col_index).value = cell_value


def extract_person_number(sheet, column_name, txt_filename):
    logging.debug(
        f"Extracting person numbers from sheet: {sheet.title}, column: {column_name}")
    column_index = None
    user_key_column_index = None
    for col in range(1, sheet.max_column + 1):
        if sheet.cell(row=1, column=col).value == column_name:
            column_index = col
        if sheet.cell(row=1, column=col).value == "User Key":
            user_key_column_index = col
    if column_index is None:
        raise Exception(
            f"Column '{column_name}' not found in the sheet '{sheet.title}'")
    case_3_numbers = []
    pull_from_otbi_found = False
    for row in range(2, sheet.max_row + 1):
        cell_value = sheet.cell(row=row, column=column_index).value
        person_number = "Unknown Format"
        if not cell_value and user_key_column_index:
            cell_value = sheet.cell(
                row=row, column=user_key_column_index).value
        if isinstance(cell_value, (int, float)):
            person_number = int(cell_value)
            case_3_numbers.append(person_number)
        elif isinstance(cell_value, str):
            if cell_value.isdigit():
                person_number = int(cell_value)
                if user_key_column_index:
                    user_key_value = sheet.cell(
                        row=row, column=user_key_column_index).value
                    if user_key_value and user_key_value.startswith("GWI Dates-"):
                        person_number = user_key_value.split("-")[1]
                        if not person_number.isdigit():
                            person_number = "Unknown Format"
                    else:
                        case_3_numbers.append(person_number)
                        person_number = "Pull from OTBI"
                        pull_from_otbi_found = True
            else:
                case_3_numbers.append(person_number)
                person_number = "Pull from OTBI"
                pull_from_otbi_found = True
        elif cell_value.startswith("WT_"):
            person_number = cell_value.split("_")[1]

        elif cell_value.startswith("PEREXTRAINFO_GWI_"):
            person_number = cell_value.split("_")[2]
        elif cell_value.startswith("SAL_") and not cell_value.startswith("SAL_ASG_"):
            matches = re.findall(r"SAL_(\d+)_\d+", cell_value)
            if matches:
                for assignment_id in matches:
                    case_3_numbers.append(assignment_id)
                person_number = "Pull from OTBI"
                pull_from_otbi_found = True
            else:
                person_number = "Unknown Format"
        elif cell_value.startswith("SAL_ASG_"):
            person_number = cell_value.split("_")[2]
        elif cell_value.endswith("-Total Progression Hours Worked"):
            person_number = cell_value.split("-")[0]
        elif cell_value.startswith("GWI Dates-"):
            person_number = cell_value.split("-")[1]
            if not person_number.isdigit():
                person_number = "Unknown Format"
            else:
                person_number = "Unknown Format"
        else:
            person_number = "Unknown Format"
        if isinstance(person_number, str):
            match = re.match(r'\d+', person_number)
            person_number = match.group(0) if match else person_number
        if isinstance(person_number, str) and person_number.isdigit():
            person_number = int(person_number)
        sheet.cell(row=row, column=1).value = person_number
        if isinstance(person_number, int):
            sheet.cell(row=row, column=1).number_format = '0'
    if pull_from_otbi_found:
        with open(txt_filename, 'w') as f:
            f.write(",\n".join(map(str, case_3_numbers)) + "\n")
        logging.debug(f"Saved person numbers to {txt_filename}")


def save_combined_workbook(combined_wb, output_dir):
    mountain_tz = pytz.timezone('US/Mountain')
    current_date = datetime.now(mountain_tz)
    day_of_creation = current_date.strftime("%A").upper()
    current_date_str = current_date.strftime("%m%d%Y")
    filename = f"WP {day_of_creation} Run {current_date_str} - HDL Fallouts.xlsx"
    save_path = os.path.join(output_dir, filename)
    combined_wb.save(save_path)
    logging.info(f"Combined workbook saved: {save_path}")
    os.startfile(save_path)


def process_dataset_folder(dataset_folder, root_dir, combined_wb):
    folder_path = os.path.join(root_dir, dataset_folder)
    if not os.path.exists(folder_path):
        logging.warning(f"Skipping non-existent folder: {dataset_folder}")
        return
    for filename in os.listdir(folder_path):
        file_path = os.path.join(folder_path, filename)
        if not filename.lower().endswith(('.xls', '.xlsx')):
            continue
        if filename.endswith('.xls'):
            new_path = convert_xls_to_xlsx(file_path)
            if not new_path:
                continue
            file_path = new_path
        try:
            wb = load_workbook(file_path)
            if 'Exported' not in wb.sheetnames:
                logging.error(f"Missing 'Exported' sheet in {filename}")
                wb.close()
                continue
            sheet = wb['Exported']
            headers = [cell.value for cell in sheet[1]]
            try:
                bo_index = headers.index('Business Object') + 1
            except ValueError:
                logging.warning(
                    f"Skipping {filename}: 'Business Object' column not found")
                wb.close()
                continue
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if row[bo_index - 1] is None:
                    business_object = 'Unknown'
                else:
                    business_object = row[bo_index - 1]
                if business_object == 'Salary':
                    sheet_name = f"{dataset_folder}Salary"
                elif business_object == 'Worker':
                    sheet_name = f"{dataset_folder}Worker"
                elif business_object == 'PersonBenefitBalance':
                    sheet_name = "ABS_EMPSYNCPSTOHCM_9"
                elif business_object == 'AssignedPayroll':
                    sheet_name = "ORCLIBAUTO_14_ASGPAY"
                else:
                    sheet_name = "NEW DATASET"
                if sheet_name not in combined_wb.sheetnames:
                    new_sheet = combined_wb.create_sheet(sheet_name)
                    new_sheet.append(headers)
                target_sheet = combined_wb[sheet_name]
                target_sheet.append(row)
            wb.close()
        except Exception as e:
            logging.error(f"Error processing {file_path}: {str(e)}")


def process_day_folder(day_folder, input_dir, output_dir):
    combined_wb = Workbook()
    if 'Sheet' in combined_wb.sheetnames:
        del combined_wb['Sheet']
    day_folder_path = os.path.join(input_dir, day_folder)
    if not os.path.exists(day_folder_path):
        logging.warning(f"Skipping non-existent folder: {day_folder}")
        return
    for subfolder in os.listdir(day_folder_path):
        subfolder_path = os.path.join(day_folder_path, subfolder)
        if not os.path.isdir(subfolder_path):
            continue
        for filename in os.listdir(subfolder_path):
            file_path = os.path.join(subfolder_path, filename)
            if not (filename.endswith('.xls') or filename.endswith('.xlsx')):
                continue
            if filename.endswith('.xls'):
                converted_path = convert_xls_to_xlsx(file_path)
                if not converted_path:
                    continue
                file_path = converted_path
                filename = os.path.basename(converted_path)
            try:
                wb = load_workbook(file_path)
                if 'Exported' not in wb.sheetnames:
                    logging.error(f"Missing 'Exported' sheet in {filename}")
                    wb.close()
                    continue
                exported_sheet = wb['Exported']
                headers = next(exported_sheet.iter_rows(
                    min_row=1, max_row=1, values_only=True))
                if 'Business Object' not in headers:
                    logging.error(
                        f"Missing 'Business Object' header in {filename}")
                    wb.close()
                    continue
                bo_index = headers.index('Business Object') + 1
                salary_sheet = None
                worker_sheet = None
                abs_49_sheet = None
                for row in exported_sheet.iter_rows(min_row=2, values_only=True):
                    if row is None:
                        logging.warning(f"No data rows in {filename}")
                        continue
                    business_object = row[bo_index -
                                          1] or "ABS_EMPSYNCPSTOHCM_9"
                    if business_object not in ['Salary', 'Worker']:
                        business_object = "ABS_EMPSYNCPSTOHCM_9"
                    if business_object == "Salary":
                        if salary_sheet is None:
                            base_sheet_name = f"{subfolder}_Salary"
                            sheet_name = base_sheet_name
                            counter = 1
                            while sheet_name in combined_wb.sheetnames:
                                sheet_name = f"{base_sheet_name}_{counter}"
                                counter += 1
                            salary_sheet = combined_wb.create_sheet(sheet_name)
                            salary_sheet.append(
                                headers)
                            salary_sheet.append(row)
                    elif business_object == "Worker":
                        if worker_sheet is None:
                            base_sheet_name = f"{subfolder}_Worker"
                            sheet_name = base_sheet_name
                            counter = 1
                            while sheet_name in combined_wb.sheetnames:
                                sheet_name = f"{base_sheet_name}_{counter}"
                                counter += 1
                            worker_sheet = combined_wb.create_sheet(sheet_name)
                            worker_sheet.append(
                                headers)
                            worker_sheet.append(row)
                    else:
                        if abs_49_sheet is None:
                            base_sheet_name = "ABS_EMPSYNCPSTOHCM_9"
                            sheet_name = base_sheet_name
                            counter = 1
                            while sheet_name in combined_wb.sheetnames:
                                sheet_name = f"{base_sheet_name}_{counter}"
                                counter += 1
                            abs_49_sheet = combined_wb.create_sheet(sheet_name)
                            abs_49_sheet.append(
                                headers)
                            abs_49_sheet.append(row)
                wb.close()
            except Exception as e:
                logging.error(f"Error processing {filename}: {e}")
                continue
    for sheet_name in combined_wb.sheetnames:
        sheet = combined_wb[sheet_name]
        check_and_update_header(sheet)
        sort_by_column(sheet, "Source System Owner")
        extract_person_number(sheet, "Source System ID", os.path.join(
            output_dir, f"{sheet_name}_Source_System_ID.txt"))
    if combined_wb.sheetnames:
        save_combined_workbook(combined_wb, output_dir)
    else:
        logging.warning(f"No valid sheets created for {day_folder}")


def main():
    configure_logging()
    try:
        root_dir = get_file_location()
        output_dir = os.path.join(
            root_dir, f"WPFallouts_{datetime.now().strftime('%Y%m%d')}")
        os.makedirs(output_dir, exist_ok=True)
        combined_wb = Workbook()
        if 'Sheet' in combined_wb.sheetnames:
            del combined_wb['Sheet']
        for folder in os.listdir(root_dir):
            if os.path.isdir(os.path.join(root_dir, folder)) and folder not in (os.path.basename(output_dir), 'build'):
                process_dataset_folder(folder, root_dir, combined_wb)
        for day_folder in ['TUESDAY', 'THURSDAY']:
            day_folder_path = os.path.join(root_dir, day_folder)
            if os.path.exists(day_folder_path) and os.path.isdir(day_folder_path):
                for subfolder in os.listdir(day_folder_path):
                    subfolder_path = os.path.join(day_folder_path, subfolder)
                    if os.path.isdir(subfolder_path):
                        process_dataset_folder(os.path.join(
                            day_folder, subfolder), root_dir, combined_wb)
        for sheet in combined_wb:
            check_and_update_header(sheet)
            sort_by_column(sheet, "Source System Owner")
            txt_path = os.path.join(
                output_dir, f"{sheet.title}_PersonNumbers.txt")
            extract_person_number(sheet, "Source System ID", txt_path)
        mountain_tz = pytz.timezone('US/Mountain')
        logging.info("All files processed.")
        current_date = datetime.now(mountain_tz)
        filename = f"WP {current_date.strftime('%A')} Run {current_date.strftime('%m%d%Y')} - HDL Fallouts.xlsx"
        output_path = os.path.join(output_dir, filename)
        combined_wb.save(output_path)
        os.startfile(output_path)
    except Exception as e:
        logging.error(f"An error occurred: {e}", exc_info=True)


if __name__ == "__main__":
    main()

# test push
